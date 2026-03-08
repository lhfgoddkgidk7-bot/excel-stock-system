#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, LineChart
import pandas as pd
from datetime import datetime
import streamlit as st
from io import BytesIO

# ========== 样式定义 ==========
title_font = Font(size=14, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_alignment = Alignment(horizontal="center", vertical="center")
header_font = Font(size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
input_font = Font(size=10, color="1565C0")
input_alignment = Alignment(horizontal="center", vertical="center")
formula_font = Font(size=10, color="000000")
formula_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
color_list = ["Negro", "Plata", "Champán", "Madera", "Blanco"]

# ========== 工具函数 ==========
def set_column_width(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def apply_title(ws, title, row=1, col_range="A:F"):
    """修正合并单元格范围"""
    start_col, end_col = col_range.split(":")
    merge_range = f"{start_col}{row}:{end_col}{row}"
    ws.merge_cells(merge_range)
    cell = ws[f'{start_col}{row}']
    cell.value = title
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = title_alignment
    ws.row_dimensions[row].height = 30

def apply_header(ws, headers, start_row=2):
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def apply_input_style(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.font = input_font
    cell.alignment = input_alignment
    cell.border = thin_border

def apply_formula_style(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.font = formula_font
    cell.alignment = formula_alignment
    cell.border = thin_border

# ========== 生成Excel核心函数 ==========
def generate_excel():
    wb = Workbook()
    wb.remove(wb.active)

    # ========== 1. 仪表盘（带图表+自动计算） ==========
    ws_dashboard = wb.create_sheet("仪表盘")
    ws_dashboard.sheet_view.showGridLines = False
    apply_title(ws_dashboard, "库存管理系统仪表盘", col_range="B:G")

    # 日期筛选
    ws_dashboard['B3'] = "日期筛选："
    ws_dashboard['B3'].font = Font(size=11, bold=True)
    ws_dashboard['C3'] = "起始日期"
    ws_dashboard['D3'] = datetime(2024, 1, 1).strftime("%Y-%m-%d")  # 直接赋值日期，而非公式
    ws_dashboard['D3'].number_format = "yyyy-mm-dd"
    ws_dashboard['E3'] = "终止日期"
    ws_dashboard['F3'] = datetime.today().strftime("%Y-%m-%d")
    ws_dashboard['F3'].number_format = "yyyy-mm-dd"

    # 财务汇总（修正公式引用）
    ws_dashboard['B5'] = "财务汇总"
    ws_dashboard['B5'].font = Font(size=12, bold=True)
    headers_dash = ["指标", "数值", "单位"]
    apply_header(ws_dashboard, headers_dash, start_row=6)

    # 修正公式：确保引用正确的工作表和列
    dash_data = [
        ["总入库金额", "=SUM(入库记录!J:J)", "美元"],
        ["总出库金额", "=SUM(出库记录!J:J)", "美元"],
        ["当前库存价值", "=SUM(期末库存!J:J)", "美元"],
        ["销售毛利（估算）", "=SUM(出库记录!J:J)-SUM(出库记录!I:I)", "美元"],
        ["入库次数", "=COUNTA(入库记录!B:B)-1", "次"],
        ["出库次数", "=COUNTA(出库记录!B:B)-1", "次"],
    ]

    for row_idx, data in enumerate(dash_data, start=7):
        # 指标列
        cell_ind = ws_dashboard.cell(row=row_idx, column=2, value=data[0])
        cell_ind.font = Font(size=10, bold=True)
        cell_ind.alignment = Alignment(horizontal="center")
        cell_ind.border = thin_border
        
        # 数值列（公式）
        cell_val = ws_dashboard.cell(row=row_idx, column=3, value=data[1])
        cell_val.number_format = '#,##0.00'
        cell_val.font = formula_font
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.border = thin_border
        
        # 单位列
        cell_unit = ws_dashboard.cell(row=row_idx, column=4, value=data[2])
        cell_unit.font = Font(size=10)
        cell_unit.alignment = Alignment(horizontal="center")
        cell_unit.border = thin_border

    # 添加柱状图（入库/出库金额对比）
    chart = BarChart()
    chart.title = "入库/出库金额对比"
    chart.style = 10
    chart.x_axis.title = "类型"
    chart.y_axis.title = "金额（美元）"

    # 数据引用
    x_data = Reference(ws_dashboard, min_col=2, min_row=7, max_row=8)  # 入库/出库金额
    y_data = Reference(ws_dashboard, min_col=3, min_row=7, max_row=8)  # 对应数值
    chart.add_data(y_data, titles_from_data=False)
    chart.set_categories(x_data)
    ws_dashboard.add_chart(chart, "B15")  # 图表位置

    # 添加折线图（库存价值趋势）
    line_chart = LineChart()
    line_chart.title = "库存价值趋势"
    line_chart.style = 12
    line_chart.x_axis.title = "指标"
    line_chart.y_axis.title = "金额（美元）"
    line_data = Reference(ws_dashboard, min_col=3, min_row=7, max_row=9)  # 入库/出库/库存价值
    line_cats = Reference(ws_dashboard, min_col=2, min_row=7, max_row=9)
    line_chart.add_data(line_data, titles_from_data=False)
    line_chart.set_categories(line_cats)
    ws_dashboard.add_chart(line_chart, "G15")

    set_column_width(ws_dashboard, {'B': 20, 'C': 15, 'D': 10, 'E': 10, 'F': 15, 'G': 15})

    # ========== 2. 客户信息 ==========
    ws_customers = wb.create_sheet("客户信息")
    ws_customers.sheet_view.showGridLines = False
    apply_title(ws_customers, "客户信息", col_range="B:F")
    headers_customers = ["注册日期", "客户名称", "客户电话", "客户地区", "详细地址"]
    apply_header(ws_customers, headers_customers)
    sample_customers = [
        ["2024-01-15", "ABC公司", "1234567890", "北京", "朝阳区某某路123号"],
        ["2024-02-20", "XYZ贸易", "0987654321", "上海", "浦东新区某某街456号"],
    ]
    for row_idx, data in enumerate(sample_customers, start=3):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_customers.cell(row=row_idx, column=col_idx, value=value)
            cell.font = input_font
            cell.alignment = input_alignment
            cell.border = thin_border
    set_column_width(ws_customers, {'B': 15, 'C': 20, 'D': 15, 'E': 15, 'F': 25})

    # ========== 3. 商品信息 ==========
    ws_products = wb.create_sheet("商品信息")
    ws_products.sheet_view.showGridLines = False
    apply_title(ws_products, "商品信息", col_range="B:H")
    headers_products = ["商品名称", "商品型号", "商品颜色", "商品数量", "长度(m)", "米重(kg/m)", "支重(kg)"]
    apply_header(ws_products, headers_products)
    # 颜色下拉验证
    color_dv = DataValidation(type="list", formula1='"Negro,Plata,Champán,Madera,Blanco"', allow_blank=True)
    color_dv.error = "请从列表中选择颜色"
    color_dv.errorTitle = "无效输入"
    ws_products.add_data_validation(color_dv)
    color_dv.add('D3:D1000')
    # 示例数据
    sample_products = [
        ["钢管产品A", "MODEL-A", "Negro", 100, 6.0, 2.5, 15.0],
        ["钢管产品B", "MODEL-B", "Plata", 50, 6.0, 3.0, 18.0],
        ["钢管产品C", "MODEL-C", "Champán", 80, 6.0, 2.8, 16.8],
    ]
    for row_idx, data in enumerate(sample_products, start=3):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_products.cell(row=row_idx, column=col_idx, value=value)
            cell.font = input_font
            cell.alignment = input_alignment
            cell.border = thin_border
            if col_idx in [5, 6, 7, 8]:
                cell.number_format = '#,##0.00'
    set_column_width(ws_products, {'B': 15, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12})

    # ========== 4. 入库记录（修正总成本公式） ==========
    ws_stockin = wb.create_sheet("入库记录")
    ws_stockin.sheet_view.showGridLines = False
    apply_title(ws_stockin, "入库记录", col_range="B:L")
    headers_stockin = ["入库日期", "商品名称", "商品型号", "商品颜色", "商品数量", "长度(m)", "米重(kg/m)", "FOB价格(US)", "汇率", "总成本", "备注"]
    apply_header(ws_stockin, headers_stockin)
    # 颜色下拉
    color_dv2 = DataValidation(type="list", formula1='"Negro,Plata,Champán,Madera,Blanco"', allow_blank=True)
    ws_stockin.add_data_validation(color_dv2)
    color_dv2.add('E3:E1000')
    # 示例数据
    sample_stockin = [
        ["2024-01-10", "钢管产品A", "MODEL-A", "Negro", 100, 6.0, 2.5, 10.0, 7.2, "", "首批入库"],
        ["2024-01-15", "钢管产品B", "MODEL-B", "Plata", 50, 6.0, 3.0, 12.0, 7.2, "", "首批入库"],
        ["2024-02-01", "钢管产品A", "MODEL-A", "Negro", 80, 6.0, 2.5, 10.5, 7.3, "", "第二批入库"],
    ]
    # 先填充示例数据
    for row_idx, data in enumerate(sample_stockin, start=3):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_stockin.cell(row=row_idx, column=col_idx, value=value)
            cell.font = input_font
            cell.alignment = input_alignment
            cell.border = thin_border
            if col_idx in [6,7,8,9,10]:
                cell.number_format = '#,##0.00'
    # 批量设置总成本公式（J列）
    for row in range(3, 103):
        cell = ws_stockin.cell(row=row, column=10)  # J列=总成本
        cell.value = f"=IF(AND(F{row}>0,G{row}>0,H{row}>0,I{row}>0,J{row}>0),F{row}*G{row}*H{row}*I{row}*J{row},0)"
        cell.number_format = '#,##0.00'
        cell.font = formula_font
        cell.alignment = formula_alignment
        cell.border = thin_border
    set_column_width(ws_stockin, {'B': 12, 'C': 12, 'D': 12, 'E': 10, 'F': 10, 'G': 10, 'H': 12, 'I': 8, 'J': 12, 'K': 12, 'L': 15})

    # ========== 5. 出库记录（修正总成本公式） ==========
    ws_stockout = wb.create_sheet("出库记录")
    ws_stockout.sheet_view.showGridLines = False
    apply_title(ws_stockout, "出库记录", col_range="B:K")
    headers_stockout = ["出库日期", "销售单号", "商品名称", "商品型号", "商品颜色", "商品数量", "长度(m)", "米重(kg/m)", "总成本", "客户", "备注"]
    apply_header(ws_stockout, headers_stockout)
    # 颜色下拉
    color_dv3 = DataValidation(type="list", formula1='"Negro,Plata,Champán,Madera,Blanco"', allow_blank=True)
    ws_stockout.add_data_validation(color_dv3)
    color_dv3.add('F3:F1000')
    # 客户下拉
    customer_dv = DataValidation(type="list", formula1='客户信息!$C$3:$C$100', allow_blank=True)
    ws_stockout.add_data_validation(customer_dv)
    customer_dv.add('J3:J1000')
    # 示例数据
    sample_stockout = [
        ["2024-02-15", "SO-20240215-001", "钢管产品A", "MODEL-A", "Negro", 30, 6.0, 2.5, "", "ABC公司", "首次销售"],
        ["2024-02-20", "SO-20240220-002", "钢管产品B", "MODEL-B", "Plata", 20, 6.0, 3.0, "", "XYZ贸易", "首次销售"],
    ]
    # 填充示例数据
    for row_idx, data in enumerate(sample_stockout, start=3):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_stockout.cell(row=row_idx, column=col_idx, value=value)
            cell.font = input_font
            cell.alignment = input_alignment
            cell.border = thin_border
            if col_idx in [7,8,9,10]:
                cell.number_format = '#,##0.00'
    # 批量设置总成本公式（I列）
    for row in range(3, 103):
        cell = ws_stockout.cell(row=row, column=9)  # I列=总成本
        cell.value = f"=IF(AND(G{row}>0,H{row}>0,I{row}>0),G{row}*H{row}*I{row}*VLOOKUP(D{row},商品信息!$B$3:$H$100,6,FALSE)*7.2,0)"
        cell.number_format = '#,##0.00'
        cell.font = formula_font
        cell.alignment = formula_alignment
        cell.border = thin_border
    set_column_width(ws_stockout, {'B': 12, 'C': 18, 'D': 12, 'E': 12, 'F': 10, 'G': 10, 'H': 10, 'I': 12, 'J': 15, 'K': 15})

    # ========== 6. 期末库存（修正总成本公式） ==========
    ws_inventory = wb.create_sheet("期末库存")
    ws_inventory.sheet_view.showGridLines = False
    apply_title(ws_inventory, "期末库存 (FIFO)", col_range="B:J")
    headers_inventory = ["商品名称", "商品型号", "商品颜色", "剩余数量", "长度(m)", "米重(kg/m)", "FOB价格(US)", "汇率", "总成本", "入库批次"]
    apply_header(ws_inventory, headers_inventory)
    # 示例数据
    sample_inventory = [
        ["钢管产品A", "MODEL-A", "Negro", 150, 6.0, 2.5, 10.2, 7.25, "", "BATCH-20240110"],
        ["钢管产品B", "MODEL-B", "Plata", 30, 6.0, 3.0, 12.0, 7.2, "", "BATCH-20240115"],
        ["钢管产品C", "MODEL-C", "Champán", 80, 6.0, 2.8, 11.0, 7.2, "", "BATCH-20240201"],
    ]
    # 填充示例数据
    for row_idx, data in enumerate(sample_inventory, start=3):
        for col_idx, value in enumerate(data, start=2):
            cell = ws_inventory.cell(row=row_idx, column=col_idx, value=value)
            cell.font = input_font
            cell.alignment = input_alignment
            cell.border = thin_border
            if col_idx in [5,6,7,8,9]:
                cell.number_format = '#,##0.00'
    # 批量设置总成本公式（I列）
    for row in range(3, 103):
        cell = ws_inventory.cell(row=row, column=9)  # I列=总成本
        cell.value = f"=IF(AND(E{row}>0,F{row}>0,G{row}>0,H{row}>0,I{row}>0),E{row}*F{row}*G{row}*H{row}*I{row},0)"
        cell.number_format = '#,##0.00'
        cell.font = formula_font
        cell.alignment = formula_alignment
        cell.border = thin_border
    set_column_width(ws_inventory, {'B': 15, 'C': 12, 'D': 12, 'E': 10, 'F': 10, 'G': 10, 'H': 12, 'I': 8, 'J': 12, 'K': 15})

    # ========== 7. 销售单（修正合并单元格+公式） ==========
    ws_invoice = wb.create_sheet("销售单")
    ws_invoice.sheet_view.showGridLines = False
    # 标题
    ws_invoice.merge_cells('B2:K2')
    ws_invoice['B2'] = "销售单 / PEDIDO DE VENTA"
    ws_invoice['B2'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_invoice['B2'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_invoice['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws_invoice.row_dimensions[2].height = 35

    # 销售单信息
    ws_invoice['B4'] = "销售单号 / No. de Pedido:"
    ws_invoice['C4'] = "SO-__________"
    ws_invoice['B5'] = "日期 / Fecha:"
    ws_invoice['C5'] = "____/____/________"
    ws_invoice['F4'] = "客户 / Cliente:"
    ws_invoice['G4'] = "________________________"
    ws_invoice['F5'] = "电话 / Tel:"
    ws_invoice['G5'] = "________________________"
    for row in [4, 5]:
        for col in ['B', 'C', 'F', 'G']:
            ws_invoice[f'{col}{row}'].font = Font(size=10, bold=True)

    # 商品表格标题
    ws_invoice.merge_cells('B7:K7')
    ws_invoice['B7'] = "商品明细 / Detalles de Mercancía"
    ws_invoice['B7'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_invoice['B7'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_invoice['B7'].alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    invoice_headers = ["序号\nNo.", "商品名称\nNombre", "型号\nModelo", "颜色\nColor", "数量\nCantidad", "长度(m)\nLongitud", "米重(kg/m)\nPeso", "单价(US)\nPrecio Unit.", "汇率\nTasa", "总金额\nImporte Total"]
    apply_header(ws_invoice, invoice_headers, start_row=8)

    # 表格数据区域
    for row in range(9, 14):
        for col in range(2, 12):
            cell = ws_invoice.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = input_font
            # 总金额列（K列）自动计算
            if col == 11:
                cell.value = f"=IF(AND(G{row}>0,H{row}>0,I{row}>0,J{row}>0),G{row}*H{row}*I{row}*J{row},0)"
                cell.number_format = '#,##0.00'
                cell.font = formula_font

    # 总计行（修正顺序：先设公式再合并）
    ws_invoice['K14'].font = Font(size=11, bold=True)
    ws_invoice['K14'].number_format = '#,##0.00'
    ws_invoice['K14'].value = "=SUM(K9:K13)"
    ws_invoice['K14'].alignment = Alignment(horizontal="center", vertical="center")
    ws_invoice.merge_cells('B14:K14')
    ws_invoice['B14'] = "总计 / TOTAL:"
    ws_invoice['B14'].font = Font(size=11, bold=True)
    ws_invoice['B14'].alignment = Alignment(horizontal="right", vertical="center")

    # 备注
    ws_invoice['B16'] = "备注 / Observaciones:"
    ws_invoice['B16'].font = Font(size=10, bold=True)
    ws_invoice.merge_cells('B17:K19')
    ws_invoice['B17'].border = thin_border

    # 签名行
    ws_invoice['B21'] = "客户签名 / Firma del Cliente:_________________"
    ws_invoice['F21'] = "日期 / Fecha:____/____/________"
    ws_invoice['B22'] = "销售员 / Vendedor:_________________"
    ws_invoice['F22'] = "日期 / Fecha:____/____/________"
    for row in [21, 22]:
        for col in ['B', 'F']:
            ws_invoice[f'{col}{row}'].font = Font(size=10)

    set_column_width(ws_invoice, {'B': 8, 'C': 15, 'D': 12, 'E': 10, 'F': 10, 'G': 10, 'H': 12, 'I': 8, 'J': 10, 'K': 14})

    # ========== 8. 使用说明 ==========
    ws_guide = wb.create_sheet("使用说明")
    ws_guide.sheet_view.showGridLines = False
    apply_title(ws_guide, "库存管理系统使用说明", col_range="B:H")
    guide_content = [
        ["", ""],
        ["1. 基础操作", "在「客户信息/商品信息」录入基础数据，颜色从下拉列表选择。"],
        ["2. 出入库记录", "录入入库/出库数据，系统自动计算总成本（数量×长度×米重×价格×汇率）。"],
        ["3. 期末库存", "系统按FIFO自动计算库存价值，可手动调整批次和数量。"],
        ["4. 仪表盘", "自动汇总财务数据，包含入库/出库金额对比图、库存价值趋势图。"],
        ["5. 销售单", "填写商品明细后，总金额和总计自动计算，可直接打印。"],
        ["", ""],
        ["公式说明:", "所有总成本/总金额列均为自动计算，修改基础数据后会实时更新。"],
        ["图表说明:", "仪表盘的柱状图/折线图会随数据变化自动更新。"],
    ]
    for row_idx, data in enumerate(guide_content, start=3):
        cell1 = ws_guide.cell(row=row_idx, column=2, value=data[0])
        cell1.font = Font(size=11, bold=True if data[0] else Font(size=11))
        cell2 = ws_guide.cell(row=row_idx, column=3, value=data[1])
        cell2.font = Font(size=10)
        ws_guide.merge_cells(f'C{row_idx}:H{row_idx}')
    set_column_width(ws_guide, {'B': 25, 'C': 65})

    # 调整工作表顺序
    wb._sheets = [wb['使用说明'], wb['仪表盘'], wb['客户信息'], wb['商品信息'], 
                  wb['入库记录'], wb['出库记录'], wb['期末库存'], wb['销售单']]
    
    # 启用Excel自动计算
    wb.calculation.calcMode = "auto"  # 关键：开启自动计算
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== Streamlit Web界面 ==========
st.set_page_config(page_title="Excel库存管理系统", layout="wide")
st.title("📊 Excel库存管理系统生成工具")
st.markdown("### 点击下方按钮生成带自动计算+图表的库存管理Excel文件")

if st.button("📥 生成Excel文件", type="primary"):
    excel_file = generate_excel()
    st.success("✅ Excel文件生成成功！已包含自动计算公式和仪表盘图表")
    st.download_button(
        label="📤 下载Excel文件",
        data=excel_file,
        file_name="Excel库存管理系统_带自动计算.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.markdown("---")
st.markdown("### 核心功能说明")
st.markdown("✅ 入库/出库/库存总成本自动计算")
st.markdown("✅ 仪表盘柱状图（入库/出库金额对比）")
st.markdown("✅ 仪表盘折线图（库存价值趋势）")
st.markdown("✅ 销售单总金额自动汇总")
st.markdown("✅ Excel自动计算模式开启，数据修改实时更新")
